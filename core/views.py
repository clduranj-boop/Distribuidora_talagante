import json
import base64
from os import link
import socket
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from django.urls import reverse
from rest_framework.views import APIView
from django.contrib.admin.views.decorators import staff_member_required
from rest_framework.response import Response
from rest_framework import status
from .models import COMUNAS_RM, DireccionEnvio, DireccionGuardada, Producto, Orden, Carrito, ItemCarrito, Perfil, ItemOrden
import urllib.parse
from .serializers import ProductoSerializer
from .forms import ProductoForm
from django.utils import timezone
from django.db.models import Q
from decimal import Decimal
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import ensure_csrf_cookie
import requests
from datetime import timedelta
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.db import transaction
from .models import Carrito, ItemCarrito, Orden, ItemOrden, User
from .serializers import OrdenSerializer
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.core.paginator import Paginator
from .models import Producto, Orden
from django.core.mail import send_mail
from django.conf import settings
from django.db import models
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import requests
from decimal import Decimal
from django.utils.crypto import get_random_string
from .models import Producto
from .forms import EscaneoEntradaForm, ProductoRapidoForm
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives
from .models import Banner





#Validación de contraseña y registro
from .validators import validar_contraseña_fuerte  
import re
#Verificacion Correo seguro
from django.core.mail import send_mail
# Decorador para verificar si el usuario es superusuario
from .models import Perfil, CodigoVerificacion
import random


def is_superuser(user):
    return user.is_superuser


#Vistas
def home(request):
    
    banners = Banner.objects.filter(activo=True).order_by('orden')

    
    productos_destacados = Producto.objects.filter(activo=True).order_by('?')[:4]
    
    context = {
        'mensaje': 'Bienvenido a Distribuidora Talagante',
        'productos': productos_destacados,
        'banners': banners  
    }
    return render(request, 'core/home.html', context)

@ensure_csrf_cookie
def login_view(request):
    if request.method == 'POST':
        identificador = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        if not identificador or not password:
            return render(request, 'core/login.html', {
                'error': 'Debes ingresar usuario/correo y contraseña'
            })

        # 1. Intentar autenticar directamente (por si es el username)
        user = authenticate(request, username=identificador, password=password)

        # 2. Si no funciona con el username, buscar por correo
        if user is None:
            try:
                usuario_por_correo = User.objects.get(email__iexact=identificador)
                user = authenticate(request, username=usuario_por_correo.username, password=password)
            except User.DoesNotExist:
                user = None

        # 3. Si encontró usuario → loguear
        if user is not None:
            if user.is_active:
                login(request, user)
                # Redirecciones según rol
                if user.is_superuser:
                    return redirect('admin_home')
                try:
                    perfil = user.perfil
                    if perfil.es_admin:
                        return redirect('admin_panel')
                    else:
                        return redirect('catalogo')
                except:
                    return redirect('catalogo')
            else:
                return render(request, 'core/login.html', {
                    'error': 'Tu cuenta aún no está verificada. Revisa tu correo.'
                })
        else:
            return render(request, 'core/login.html', {
                'error': 'Usuario/correo o contraseña incorrectos'
            })

    return render(request, 'core/login.html')


def verificar_codigo(request):
    # FORZAR QUE EL USUARIO ESTÉ LOGUEADO (aunque is_active=False)
    if not request.user.is_authenticated:
        # Buscar si hay un usuario reciente con código pendiente
        try:
            ultimo_codigo = CodigoVerificacion.objects.filter(
                creado_en__gte=timezone.now() - timedelta(minutes=15)
            ).latest('creado_en')
            user = ultimo_codigo.usuario
            # LOGUEAR FORZADAMENTE AL USUARIO
            login(request, user)
        except CodigoVerificacion.DoesNotExist:
            return redirect('login')

    user = request.user

    # Si ya está activo → ir al home
    if user.is_active:
        return redirect('home')

    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        try:
            cod_obj = CodigoVerificacion.objects.get(
                usuario=user,
                codigo=codigo,
                expirado=False
            )
            if cod_obj.creado_en >= timezone.now() - timedelta(minutes=10):
                user.is_active = True
                user.save()
                cod_obj.expirado = True
                cod_obj.save()
                messages.success(request, "¡Cuenta verificada con éxito! Bienvenido.")
                return redirect('home')
            else:
                messages.error(request, "El código ha expirado.")
        except CodigoVerificacion.DoesNotExist:
            messages.error(request, "Código incorrecto.")

    return render(request, 'core/verificar_codigo.html')


# Lista de dominios temporales
DOMINIOS_TEMPORALES = {
    '10minutemail.com', 'tempmail.org', 'guerrillamail.com', 'mailinator.com',
    'yopmail.com', 'disposablemail.com', 'throwawaymail.com', 'sharklasers.com',
    'maildrop.cc', 'getnada.com', 'armyspy.com', 'cuvox.de', 'dayrep.com',
    'einrot.com', 'fleckens.hu', 'gustr.com', 'jourrapide.com', 'rhyta.com',
    'superrito.com', 'teleworm.us', 'trashmail.com', 'wegwerfmail.de',
    'mail.tm', 'mintemail.com', 'temp-mail.org', 'tmpmail.org'
}

def validar_correo_real(email: str):
    email = email.strip().lower()

    # 1. Formato básico
    if not re.match(r"^[\w\.\+\-]+\@[\w]+\.[a-z]{2,}$", email):
        raise ValidationError("Correo con formato inválido.")

    # 2. Extraer dominio
    try:
        dominio = email.split('@')[1]
    except IndexError:
        raise ValidationError("Correo inválido.")

    # 3. Bloquear temporales
    if dominio in DOMINIOS_TEMPORALES:
        raise ValidationError("No se permiten correos temporales.")

    # 4. Validación simple: intentar conectar al puerto 25 (SMTP)
    # Esto funciona en el 98% de los casos sin instalar nada
    try:
        socket.create_connection((dominio, 25), timeout=5)
        return True
    except (socket.gaierror, socket.timeout, OSError):
        # Si falla, es muy probable que el dominio no exista o no acepte correo
        raise ValidationError("Este correo no parece real. Usa uno válido (Gmail, Hotmail, empresa, etc.).")

def register(request):
    datos_form = request.POST if request.method == "POST" else {}

    if request.method == 'POST':
        # --- Capturar datos ---
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        nombre = request.POST.get('nombre', '').strip().title()
        apellido_paterno = request.POST.get('apellido_paterno', '').strip().title()
        apellido_materno = request.POST.get('apellido_materno', '').strip().title()
        rut_raw = request.POST.get('rut', '').strip().upper()
        telefono = request.POST.get('telefono', '').strip()

        # ==================== LIMPIAR RUT ====================
        rut_sin_formato = re.sub(r'[^\dKk]', '', rut_raw).upper()
        if len(rut_sin_formato) < 8:
            messages.error(request, "El RUT debe tener al menos 8 dígitos.")
            return render(request, 'core/login.html', {'datos_form': datos_form, 'tab': 'register'})

        

        rut = rut_sin_formato[:-1] + '-' + rut_sin_formato[-1]

        # ==================== VALIDACIONES ====================
        errores = False

        
        
        if not all([username, email, password1, password2, nombre, apellido_paterno, rut]):
            messages.error(request, "Todos los campos obligatorios deben estar completos.")
            errores = True

        if not re.fullmatch(r'[a-zA-Z0-9]+', username):
            messages.error(request, "El usuario solo puede contener letras y números.")
            errores = True

        if User.objects.filter(username__iexact=username).exists():
            messages.error(request, "Ese usuario ya está en uso.")
            errores = True

        if User.objects.filter(email__iexact=email).exists():
            messages.error(request, "Ya existe una cuenta con ese correo.")
            errores = True

        if Perfil.objects.filter(rut=rut).exists():
            messages.error(request, "Este RUT ya está registrado.")
            errores = True

        if password1 != password2:
            messages.error(request, "Las contraseñas no coinciden.")
            errores = True

        if len(password1) < 8:
            messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
            errores = True

        if not any(c.isupper() for c in password1):
            messages.error(request, "La contraseña debe tener al menos una mayúscula.")
            errores = True

        if not any(c.isdigit() for c in password1):
            messages.error(request, "La contraseña debe tener al menos un número.")
            errores = True

        # Si hay errores → devolvemos el formulario con los datos
        if errores:
            return render(request, 'core/login.html', {
                'datos_form': datos_form,
                'tab': 'register'
            })

        # ==================== CREAR USUARIO ====================
        try:
            user = User.objects.create_user(username=username, email=email, password=password1)
            user.is_active = False
            user.save()

            login(request, user)

            Perfil.objects.create(
                usuario=user,
                nombre=nombre,
                apellido_paterno=apellido_paterno,
                apellido_materno=apellido_materno or None,
                rut=rut,
                telefono=telefono or None,
            )

            # Generar y enviar código
            codigo_obj = CodigoVerificacion(usuario=user)
            codigo_obj.codigo = f"{random.randint(100000, 999999)}"
            codigo_obj.save()

            try:
                send_mail(
                    "Código de verificación - Distribuidora Talagante",
                    f"Hola {username}!\n\nTu código es: {codigo_obj.codigo}\n\nVálido 10 minutos.",
                    'Distribuidora Talagante <no-reply@distribuidoratoralagante.cl>',
                    [email],
                    fail_silently=False,
                )
                messages.success(request, f"¡Código enviado a {email}!")
            except:
                messages.warning(request, f"Cuenta creada. Tu código es: {codigo_obj.codigo}")

            return redirect('verificar_codigo')

        except Exception as e:
            messages.error(request, "Error inesperado. Inténtalo de nuevo.")
            return render(request, 'core/login.html', {'datos_form': datos_form, 'tab': 'register'})

    # GET normal
    return render(request, 'core/login.html', {'tab': 'register'})


def catalogo(request):
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    categorias = Producto.objects.filter(activo=True).values_list('categoria', flat=True).distinct()
    if request.GET.get('categoria'):
        productos = productos.filter(categoria=request.GET.get('categoria'))
    if request.GET.get('precio_max'):
        productos = productos.filter(precio__lte=request.GET.get('precio_max'))
    context = {'productos': productos, 'categorias': categorias}
    return render(request, 'core/catalogo.html', context)

@login_required
def carrito(request):
    carrito = Carrito.objects.filter(
        usuario=request.user,
        creado__gte=timezone.now() - timezone.timedelta(minutes=15)
    ).first()

    if not carrito or not carrito.itemcarrito_set.exists():
        return render(request, 'core/carrito.html', {'mensaje': 'Tu carrito está vacío'})

    items = carrito.itemcarrito_set.all()
    total = sum(item.cantidad * item.producto.precio for item in items)

    context = {
        'items': items,
        'total': total,
    }
    return render(request, 'core/carrito.html', context)


ZONAS_TALAGANTE = ['Talagante', 'Peñaflor', 'Isla de Maipo', 'El Monte', 'Padre Hurtado']


@login_required
def checkout(request):
    carrito = Carrito.objects.filter(usuario=request.user).order_by('-creado').first()
    if not carrito or not carrito.itemcarrito_set.exists():
        messages.error(request, "Tu carrito está vacío.")
        return redirect('carrito')

    items = carrito.itemcarrito_set.all()
    # Calculamos subtotal preliminar
    subtotal = sum(item.cantidad * item.producto.precio for item in items)

    if request.method == 'POST':
        metodo_pago = request.POST.get('metodo_pago', 'transferencia')
        mensaje = request.POST.get('mensaje', '').strip()
        metodo_envio = request.POST.get('metodo_envio', 'retiro')
        comuna = request.POST.get('comuna', '') if metodo_envio == 'domicilio' else None

        # Validación de comuna
        if metodo_envio == 'domicilio' and not comuna:
            messages.error(request, "Debes seleccionar una comuna para el envío.")
            return redirect('checkout')

        # === 1. PROCESAMIENTO DE IMAGEN A BASE64 ===
        comprobante_img = request.FILES.get('comprobante')
        imagen_b64_final = None

        if comprobante_img:
            # Validar tamaño (5MB)
            if comprobante_img.size > 5 * 1024 * 1024:
                messages.error(request, "El archivo es muy pesado (Máx 5MB).")
                return redirect('checkout')
            
            # Validar tipo
            if comprobante_img.content_type not in ['image/jpeg', 'image/png', 'image/jpg']:
                messages.error(request, "Formato inválido. Solo JPG o PNG.")
                return redirect('checkout')

            # Convertir
            try:
                imagen_binaria = comprobante_img.read()
                imagen_b64_final = base64.b64encode(imagen_binaria).decode('utf-8')
            except Exception:
                messages.error(request, "Error procesando la imagen.")
                return redirect('checkout')

        # === 2. TRANSACCIÓN SEGURA (STOCK) ===
        try:
            with transaction.atomic():
                # Crear orden inicial
                orden = Orden.objects.create(
                    usuario=request.user,
                    total=0, # Se recalcula abajo
                    metodo_pago=metodo_pago,
                    mensaje_cliente=mensaje,
                    estado='confirmacion' if imagen_b64_final else 'pendiente',
                    comprobante_b64=imagen_b64_final  # <--- GUARDAMOS EL TEXTO
                )

                total_real = 0

                # Bucle seguro de stock
                for item in items:
                    # BLOQUEO DE BASE DE DATOS (select_for_update)
                    try:
                        producto_db = Producto.objects.select_for_update().get(id=item.producto.id)
                    except Producto.DoesNotExist:
                        raise ValueError(f"El producto {item.producto.nombre} ya no está disponible.")

                    if item.cantidad > producto_db.stock:
                        raise ValueError(f"No hay suficiente stock de {producto_db.nombre}")

                    # Crear ItemOrden
                    ItemOrden.objects.create(
                        orden=orden,
                        producto=producto_db,
                        cantidad=item.cantidad,
                        precio=producto_db.precio
                    )

                    # Descontar stock REAL
                    producto_db.stock -= item.cantidad
                    producto_db.save()

                    total_real += (producto_db.precio * item.cantidad)

                # === 3. LOGICA DE ENVÍO ===
                costo_envio = Decimal('0')
                
                if metodo_envio == 'domicilio':
                    DireccionEnvio.objects.create(
                        orden=orden, metodo='domicilio', comuna=comuna,
                        notas="Cliente coordinará por WhatsApp"
                    )
                    # Tu lógica de precios de envío
                    ZONAS_TALAGANTE = ['talagante', 'penaflor', 'isla_de_maipo', 'el_monte', 'padre_hurtado']
                    comuna_lower = comuna.lower().replace(' ', '_') # Normalizar un poco

                    if comuna_lower in ZONAS_TALAGANTE:
                         costo_envio = Decimal('2500')
                    elif total_real >= Decimal('40000'):
                         costo_envio = Decimal('0')
                    else:
                         costo_envio = Decimal('4500')
                else:
                    DireccionEnvio.objects.create(orden=orden, metodo='retiro')

                orden.total = total_real + costo_envio
                orden.save()

                # Limpiar carrito
                carrito.itemcarrito_set.all().delete()

            # === FIN DE LA TRANSACCIÓN ===
            # Si llegamos aquí, la orden se creó bien.

            # === 4. ENVÍO DE CORREOS ===
            try:
                nombre_cliente = (orden.usuario.perfil.nombre_completo()
                                 if hasattr(orden.usuario, 'perfil') and orden.usuario.perfil
                                 else orden.usuario.username)

                # Correo Cliente
                html_cliente = render_to_string('emails/confirmacion_cliente.html', {
                    'orden': orden, 'cliente': nombre_cliente, 'items': orden.itemorden_set.all()
                })
                email_cliente = EmailMultiAlternatives(
                    subject=f"¡Gracias por tu pedido #{orden.id}!",
                    body="Tu pedido ha sido recibido.",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[orden.usuario.email],
                )
                email_cliente.attach_alternative(html_cliente, "text/html")
                email_cliente.send()

                # Correo Admin
                html_admin = render_to_string('emails/nueva_orden_admin.html', {
                    'orden': orden, 'cliente': nombre_cliente, 'items': orden.itemorden_set.all()
                })
                email_admin = EmailMultiAlternatives(
                    subject=f"NUEVO PEDIDO #{orden.id}",
                    body=f"Nuevo pedido de {orden.usuario.email}",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[settings.ADMIN_EMAIL_RECEIVER], # Usar variable de settings mejor
                )
                email_admin.attach_alternative(html_admin, "text/html")
                
                # NOTA: Ya no adjuntamos el archivo físico porque está en Base64 en la BD.
                # El admin debe entrar al panel para ver la imagen.
                
                email_admin.send()

            except Exception as e:
                print(f"ERROR CORREOS: {e}")

            messages.success(request, f"¡Orden #{orden.id} creada con éxito!")
            return redirect('orden_exitosa', orden_id=orden.id)

        except ValueError as e:
            # Error de stock
            messages.error(request, str(e))
            return redirect('checkout')
        except Exception as e:
            # Error inesperado
            print(f"ERROR CHECKOUT: {e}")
            messages.error(request, "Ocurrió un error al procesar tu pedido.")
            return redirect('checkout')

    context = {
        'items': items,
        'total': subtotal,
        'comunas_rm': COMUNAS_RM,
    }
    return render(request, 'core/checkout.html', context)




def cambiar_password_view(request, token):
    try:
        perfil = Perfil.objects.get(temp_token=token)
        if perfil.token_expira and timezone.now() > perfil.token_expira:
            messages.error(request, "El enlace ha expirado.")
            return redirect('login')
    except Perfil.DoesNotExist:
        messages.error(request, "Enlace inválido.")
        return redirect('login')

    if request.method == "POST":
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            messages.error(request, "Las contraseñas no coinciden.")
        elif len(password1) < 8:
            messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
        else:
            perfil.usuario.set_password(password1)
            perfil.usuario.save()
            # Limpiar token
            perfil.temp_token = None
            perfil.token_expira = None
            perfil.save()
            messages.success(request, "¡Contraseña cambiada con éxito! Ya puedes iniciar sesión.")
            return redirect('login')

    return render(request, 'core/cambiar_password.html')

def recuperar_password(request):
    if request.method == "POST":
        entrada = request.POST.get('email_o_usuario', '').strip()

        if not entrada:
            messages.error(request, "Ingresa tu usuario, correo o RUT.")
            return redirect('login')

        user = None

        # 1. BUSCAR POR USUARIO
        if User.objects.filter(username__iexact=entrada).exists():
            user = User.objects.get(username__iexact=entrada)

        # 2. BUSCAR POR CORREO
        elif User.objects.filter(email__iexact=entrada).exists():
            user = User.objects.get(email__iexact=entrada)

        # 3. BUSCAR POR RUT 
        else:
            rut_limpio = re.sub(r'[^\dKk]', '', entrada).upper()
            if len(rut_limpio) >= 8:
                # Formateamos exactamente como está guardado en la BD: 12345678-k
                rut_formateado = rut_limpio[:-1] + '-' + rut_limpio[-1].lower()
                try:
                    perfil = Perfil.objects.get(rut=rut_formateado)
                    user = perfil.usuario
                except Perfil.DoesNotExist:
                    pass

        
        if user is None:
            messages.error(request, "No encontramos ninguna cuenta con ese usuario, correo o RUT.")
            return redirect('login')

        # === GENERAR TOKEN ===
        token = get_random_string(50)
        perfil, _ = Perfil.objects.get_or_create(usuario=user)
        perfil.temp_token = token
        perfil.token_expira = timezone.now() + timedelta(hours=2)
        perfil.save()

        link = request.build_absolute_uri(reverse('cambiar_password', args=[token]))

        # === CORREO FUNCIONAL EN CELULAR ===
        html_content = render_to_string('emails/recuperar_password.html', {
            'cliente': user.perfil.nombre_completo() if hasattr(user, 'perfil') and user.perfil.nombre else user.username,
            'link': link,
        })

        email = EmailMultiAlternatives(
            subject="Recuperar contraseña - Distribuidora Talagante",
            body="",
            from_email="Distribuidora Talagante <no-reply@distribuidoratoralagante.cl>",
            to=[user.email],
        )
        email.attach_alternative(html_content, "text/html")
        email.send()

        messages.success(request, f"¡Enlace enviado a {user.email}!")
        return redirect('login')

    return redirect('login')

def cambiar_correo_registro(request):
    if request.method == "POST":
        identificador = request.POST.get('identificador', '').strip()
        nuevo_correo = request.POST.get('nuevo_correo', '').strip().lower()

        if not identificador:
            messages.error(request, "Ingresa tu usuario o RUT.")
            return redirect('cambiar_correo_registro')

        if not nuevo_correo or '@' not in nuevo_correo:
            messages.error(request, "Ingresa un correo válido.")
            return redirect('cambiar_correo_registro')

        if User.objects.filter(email__iexact=nuevo_correo).exists():
            messages.error(request, "Ese correo ya está registrado en otra cuenta.")
            return redirect('cambiar_correo_registro')

        user = None

        # 1. BUSCAR POR USUARIO
        if User.objects.filter(username__iexact=identificador).exists():
            user = User.objects.get(username__iexact=identificador)

        # 2. BUSCAR POR RUT 
        else:
            rut_limpio = re.sub(r'[^\dKk]', '', identificador).upper()
            if len(rut_limpio) >= 8:
                rut_formateado = rut_limpio[:-1] + '-' + rut_limpio[-1].lower()
                try:
                    perfil = Perfil.objects.get(rut=rut_formateado)
                    user = perfil.usuario
                except Perfil.DoesNotExist:
                    pass

        if not user:
            messages.error(request, "No encontramos ninguna cuenta con ese usuario o RUT.")
            return redirect('cambiar_correo_registro')

        # === CAMBIAR CORREO Y FORZAR VERIFICACIÓN ===
        user.email = nuevo_correo
        user.is_active = False
        user.save()

        # Limpiar códigos viejos
        CodigoVerificacion.objects.filter(usuario=user).delete()

        # Generar nuevo código
        codigo_obj = CodigoVerificacion(usuario=user)
        codigo_obj.codigo = f"{random.randint(100000, 999999)}"
        codigo_obj.save()

        # Enviar al NUEVO correo
        try:
            send_mail(
                'Verifica tu nuevo correo - Distribuidora Talagante',
                f'Hola {user.username}!\n\n'
                f'Acabas de cambiar tu correo a: {nuevo_correo}\n\n'
                f'Tu código de verificación es:\n\n{codigo_obj.codigo}\n\n'
                f'Válido por 10 minutos.\n\n¡Gracias por confiar en nosotros!',
                'Distribuidora Talagante <no-reply@distribuidoratoralagante.cl>',
                [nuevo_correo],
                fail_silently=False,
            )
            messages.success(request, f"¡Código enviado al nuevo correo: {nuevo_correo}")
        except:
            messages.warning(request, f"Código generado: {codigo_obj.codigo} (no se pudo enviar correo)")

        # Logueamos para que pueda verificar
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        return redirect('verificar_codigo')

    return render(request, 'core/cambiar_correo_registro.html')

def reenviar_codigo(request):
    user_id = request.session.get('_auth_user_id')
    if not user_id:
        messages.error(request, "Sesión expirada. Inicia sesión de nuevo.")
        return redirect('login')

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return redirect('login')

    if user.is_active:
        return redirect('home')

    codigo_obj, created = CodigoVerificacion.objects.get_or_create(usuario=user)
    codigo_obj.codigo = f"{random.randint(100000, 999999)}"
    codigo_obj.creado_en = timezone.now()
    codigo_obj.expirado = False
    codigo_obj.save()

    # 4. Enviamos el correo
    try:
        send_mail(
            'Nuevo código de verificación',
            f'Hola {user.username}!\n\nTu nuevo código es:\n\n{codigo_obj.codigo}\n\nVálido por 10 minutos.',
            settings.EMAIL_HOST_USER or 'no-reply@distribuidoratoralagante.cl',
            [user.email],
            fail_silently=False,
        )
        messages.success(request, '¡Código reenviado con éxito! Revisa tu correo.')
    except Exception as e:
        messages.warning(request, f'Código generado: {codigo_obj.codigo} (no se pudo enviar correo)')

    return redirect('verificar_codigo')

@csrf_exempt
def autocompletar_direccion(request):
    if request.method != "POST":
        return JsonResponse({"sugerencias": []}, safe=False)

    try:
        data = json.loads(request.body)
        texto = data.get("q", "").strip()
        comuna = data.get("comuna", "").strip()

        if len(texto) < 2 or not comuna:
            return JsonResponse({"sugerencias": []}, safe=False)

        # User-Agent + URL correcta
        headers = {
            "User-Agent": "DistribuidoraTalagante/1.0 (+56912345678)",  # PON TU NÚMERO AQUÍ
            "Accept-Language": "es"
        }

        url = "https://nominatim.openstreetmap.org/search"
        params = {
            "q": f"{texto}, {comuna}, Santiago, Chile",
            "format": "json",
            "limit": 10,
            "countrycodes": "cl",
            "addressdetails": 1,
            "namedetails": 1
        }

        response = requests.get(url, params=params, headers=headers, timeout=15)
        
        if response.status_code != 200:
            return JsonResponse({"sugerencias": []}, safe=False)

        resultados = response.json()
        sugerencias = []

        for item in resultados:
            nombre = item.get("display_name", "")
            if "," in nombre:
                calle = nombre.split(",")[0].strip()
                if calle and len(calle) > 5 and texto.lower() in calle.lower():
                    sugerencias.append({
                        "value": calle,
                        "label": calle
                    })

        return JsonResponse({"sugerencias": sugerencias[:8]}, safe=False)

    except Exception as e:
        print("Error en autocompletar:", e) 
        return JsonResponse({"sugerencias": []}, safe=False)

@login_required
def orden_exitosa(request, orden_id):
    orden = get_object_or_404(Orden, id=orden_id, usuario=request.user)
    return render(request, 'core/orden_exitosa.html', {'orden': orden})

@login_required
def mis_compras(request):
    ordenes = Orden.objects.filter(usuario=request.user).order_by('-fecha')

    busqueda = request.GET.get('q')      
    estado_filtro = request.GET.get('estado')

    if busqueda:
        busqueda = busqueda.strip()
        if busqueda.isdigit():
            ordenes = ordenes.filter(id=busqueda)
        elif busqueda.startswith('#') and busqueda[1:].isdigit():
            ordenes = ordenes.filter(id=busqueda[1:])
    
    # 4. FILTRAR POR ESTADO
    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    # 5. Calcular subtotales
    for orden in ordenes:
        for item in orden.itemorden_set.all():
            item.subtotal_temp = item.cantidad * item.precio

    context = {
        'ordenes': ordenes,
        'ESTADOS': Orden.ESTADOS, 
    }
    return render(request, 'core/mis_compras.html', context)

def logout_view(request):
    logout(request)
    return redirect('home')

@login_required
@user_passes_test(is_superuser)
def admin_panel(request):
    estado_filtro = request.GET.get('estado', '')

    ordenes = Orden.objects.all().select_related('usuario__perfil').prefetch_related('itemorden_set__producto').order_by('-fecha')

    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)
    else:
        ordenes = ordenes.exclude(estado__in=['cancelado', 'completado'])

    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(ordenes, 25)
    page = request.GET.get('page')
    ordenes = paginator.get_page(page)

    context = {
        'ordenes': ordenes,
        'estado_filtro': estado_filtro,
    }
    return render(request, 'core/admin_panel.html', context)
    
@login_required
@user_passes_test(is_superuser)
def cambiar_estado_pedido(request, pk):
    orden = get_object_or_404(Orden, pk=pk)
    
    if request.method == "POST":
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado not in dict(Orden.ESTADOS):
            messages.error(request, "Estado inválido")
            return redirect('admin_panel')
        
        estado_anterior = orden.get_estado_display()
        orden.estado = nuevo_estado
        orden.save()

        # === ENVÍO DE CORREO BONITO AL CLIENTE ===
        if orden.usuario.email:
            try:
                from django.template.loader import render_to_string
                from django.core.mail import EmailMultiAlternatives
                from django.conf import settings
                import urllib.parse

                nombre_cliente = (orden.usuario.perfil.nombre_completo() 
                                if hasattr(orden.usuario, 'perfil') and orden.usuario.perfil 
                                else orden.usuario.username)

                mensaje_wa = f"Hola! Mi pedido es el #{orden.id} - Estado actual: {orden.get_estado_display()}"
                whatsapp_link = f"https://wa.me/56949071013?text={urllib.parse.quote(mensaje_wa)}"

                html_content = render_to_string('emails/cambio_estado.html', {
                    'cliente': nombre_cliente,
                    'pedido_id': orden.id,
                    'estado_anterior': estado_anterior,
                    'nuevo_estado': orden.get_estado_display(),
                    'total': orden.total,
                    'whatsapp_link': whatsapp_link,
                    'items': orden.itemorden_set.all(),
                })

                email = EmailMultiAlternatives(
                    subject=f"Tu pedido #{orden.id} cambió de estado",
                    body=f"Tu pedido #{orden.id} ahora está en: {orden.get_estado_display()}",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[orden.usuario.email],
                )
                email.attach_alternative(html_content, "text/html")
                email.send()

            except Exception as e:
                print(f"Error enviando correo cambio estado #{orden.id}: {e}")

        messages.success(request, f"Estado cambiado a {orden.get_estado_display()}")
        return redirect('admin_panel')
    
    return redirect('admin_panel')
    
@login_required
def actualizar_cantidad_carrito(request, item_id):
    item = get_object_or_404(ItemCarrito, id=item_id, carrito__usuario=request.user)
    
    if request.method == "POST":
        accion = request.POST.get("accion")
        try:
            nueva_cantidad = int(request.POST.get("cantidad", item.cantidad))
        except:
            nueva_cantidad = item.cantidad

        if accion == "sumar":
            if item.cantidad < item.producto.stock:
                item.cantidad += 1
        elif accion == "restar":
            if item.cantidad > 1:
                item.cantidad -= 1
        else:
            if 1 <= nueva_cantidad <= item.producto.stock:
                item.cantidad = nueva_cantidad

        item.save()
        
    
    return redirect('carrito')


from django.db.models import Sum, Count

@login_required
@user_passes_test(is_superuser)
def gestion_pedidos(request):
    estado_filtro = request.GET.get('estado', '')

    # Órdenes pendientes para la tabla principal
    qs = Orden.objects.select_related('usuario', 'usuario__perfil') \
        .prefetch_related('itemorden_set__producto') \
        .order_by('-fecha')

    if estado_filtro and estado_filtro in dict(Orden.ESTADOS):
        qs = qs.filter(estado=estado_filtro)
    else:
        qs = qs.exclude(estado__in=['cancelado', 'completado'])

    paginator = Paginator(qs, 25)
    page = request.GET.get('page')
    ordenes = paginator.get_page(page)

    # === NUEVO: VENTAS DIARIAS DETALLADAS (últimos 30 días) ===
    todas_ordenes = Orden.objects.all().order_by('-fecha')

    ventas_por_dia = {}
    for orden in todas_ordenes:
        dia = orden.fecha.date()
        if dia not in ventas_por_dia:
            ventas_por_dia[dia] = []
        ventas_por_dia[dia].append(orden)

    # Ordenar días y limitar a últimos 30
    dias_ordenados = sorted(ventas_por_dia.keys(), reverse=True)
    if len(dias_ordenados) > 30:
        dias_ordenados = dias_ordenados[:30]

    ventas_diarias_agrupadas = [(dia, ventas_por_dia[dia]) for dia in dias_ordenados]

    # === NUEVO: RESUMEN MENSUAL (últimos 12 meses) ===
    ventas_mensuales = todas_ordenes.extra({
        'mes': "strftime('%%m', fecha)",
        'año': "strftime('%%Y', fecha)"
    }).values('mes', 'año').annotate(
        total_ventas=Sum('total'),
        num_pedidos=Count('id')
    ).order_by('-año', '-mes')[:12]

    # Formatear nombre del mes (opcional, para que se vea bonito)
    meses_nombres = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
    }
    for item in ventas_mensuales:
        item['nombre_mes'] = meses_nombres.get(item['mes'], item['mes'])

    context = {
        'ordenes': ordenes,
        'estado_filtro': estado_filtro,
        'ventas_diarias_agrupadas': ventas_diarias_agrupadas,  # Para el detalle por día
        'ventas_mensuales': ventas_mensuales,                 # Para el resumen mensual
    }
    return render(request, 'core/gestion_pedidos.html', context)


@login_required
def add_to_carrito(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id, activo=True)
    
    if producto.stock <= 0:
        messages.error(request, f"¡{producto.nombre} está sin stock!")
        return redirect('catalogo')

    
    carrito, _ = Carrito.objects.get_or_create(
        usuario=request.user,
        creado__gte=timezone.now() - timedelta(minutes=15)
    )

    if request.method == "POST":
        try:
            cantidad = int(request.POST.get("cantidad", 1))
            if cantidad < 1:
                cantidad = 1
            if cantidad > producto.stock:
                messages.error(request, f"Solo hay {producto.stock} unidades disponibles de {producto.nombre}")
                return redirect('catalogo')
        except:
            cantidad = 1
    else:
        cantidad = 1

    item, created = ItemCarrito.objects.get_or_create(carrito=carrito, producto=producto)
    
    if created:
        item.cantidad = cantidad
    else:
        nuevo_total = item.cantidad + cantidad
        if nuevo_total > producto.stock:
            messages.warning(request, f"Solo se añadieron las unidades disponibles de {producto.nombre}")
            item.cantidad = producto.stock
        else:
            item.cantidad = nuevo_total

    item.save()

    messages.success(
        request,
        f'<strong>{producto.nombre}</strong> × {cantidad} añadido al carrito '
        f'<span class="badge bg-success ms-2">{item.cantidad} und</span>',
        extra_tags='safe'
    )
    return redirect('catalogo')

from django.shortcuts import render, redirect
from .models import ItemOrden, Producto # 


@login_required
def remove_from_carrito(request, item_id):
    
    item = get_object_or_404(ItemCarrito, id=item_id, carrito__usuario=request.user)
    
    item.delete()
    
    return redirect('carrito')


class ProductoListAPIView(APIView):
    def get(self, request):
        productos = Producto.objects.filter(activo=True).order_by('nombre')
        categoria = request.query_params.get('categoria')
        precio_max = request.query_params.get('precio_max')
        if categoria:
            productos = productos.filter(categoria=categoria)
        if precio_max:
            try:
                productos = productos.filter(precio__lte=float(precio_max))
            except ValueError:
                return Response({"error": "precio_max debe ser un número válido"}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ProductoSerializer(productos, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@login_required
@user_passes_test(is_superuser)
def admin_home(request):
    productos_bajo_stock = Producto.objects.filter(stock__lt=10)
    estado_filtro = request.GET.get('estado', '')  # Filtro por estado desde la URL
    ordenes_query = Orden.objects.filter(estado__in=['pendiente', 'confirmacion', 'preparacion'])
    
    if estado_filtro:
        ordenes_query = ordenes_query.filter(estado=estado_filtro)
    
    ordenes = ordenes_query.order_by('-fecha')
    paginator = Paginator(ordenes, 10)  # 10 órdenes por página
    page_number = request.GET.get('page')
    ordenes_paginated = paginator.get_page(page_number)

    item_id = request.GET.get('delete_item')
    orden_id = request.GET.get('delete_orden')

    if item_id:
        try:
            item = ItemOrden.objects.get(id=item_id)
            orden = item.orden
            producto = item.producto
            cantidad = item.cantidad
            precio = item.precio

            # 1. BORRAMOS EL ÍTEM
            # Al ejecutar esta línea, la "Señal" en models.py detecta el borrado 
            # y devuelve el stock automáticamente.
            item.delete() 

            # --- CORRECCIÓN APLICADA ---
            # Borramos estas líneas para evitar que se sume dos veces.
            # producto.stock += cantidad  <-- ELIMINADO
            # producto.save()             <-- ELIMINADO
            # ---------------------------

            # 2. Actualizamos el precio total de la orden
            orden.total -= cantidad * precio
            orden.save()

            messages.success(request, f'Ítem "{producto.nombre}" eliminado de la orden #{orden.id}. Stock restaurado.')
        except ItemOrden.DoesNotExist:
            messages.error(request, 'Ítem no encontrado.')

    elif orden_id:
        try:
            orden = Orden.objects.get(id=orden_id)
            orden.estado = 'cancelado'  # Esto activa el metodo save para restaurar stock
            orden.save()
            messages.success(request, f'Orden #{orden.id} cancelada y stock restaurado.')
        except Orden.DoesNotExist:
            messages.error(request, 'Orden no encontrada.')
            
    context = {
        'productos_bajo_stock': productos_bajo_stock,
        'ordenes': ordenes_paginated,
        'estado_filtro': estado_filtro,
    }
    return render(request, 'core/admin_home.html', context)

@login_required
@user_passes_test(is_superuser)
def producto_list(request):
    productos = Producto.objects.all().order_by('nombre')
    return render(request, 'core/admin_productos.html', {'productos': productos})

@login_required
@user_passes_test(is_superuser)
def producto_create(request):
    codigo_prellenado = request.GET.get('codigo_barras', '')

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Producto creado correctamente")
            return redirect('producto_list')
    else:
        initial_data = {}
        if codigo_prellenado:
            initial_data['codigo_barras'] = codigo_prellenado
        form = ProductoForm(initial=initial_data)

    return render(request, 'core/producto_form.html', {
        'form': form,
        'action': 'Crear',
        'codigo_prellenado': codigo_prellenado  
    })

@login_required
@user_passes_test(is_superuser)
def producto_update(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, "Producto actualizado correctamente")
            return redirect('producto_list')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'core/producto_form.html', {'form': form, 'action': 'Editar', 'producto': producto})

@login_required
@user_passes_test(is_superuser)
def producto_delete(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        producto.delete()
        messages.success(request, 'Producto eliminado exitosamente.')
        return redirect('producto_list')
    return render(request, 'core/producto_confirm_delete.html', {'producto': producto})

@login_required
@user_passes_test(is_superuser)  
def orden_detail(request, pk):
    orden = get_object_or_404(Orden, pk=pk)
    return render(request, 'core/orden_detail.html', {'orden': orden})

def test_endpoint_view(request):
    response_data = None
    if request.method == 'POST':
        from django.test import Client
        try:
            client = Client()
            response = client.post(reverse('probar_orden'))
            if response.status_code == 201:
                response_data = response.json()
            else:
                response_data = {"error": "Fallo en la solicitud", "detalle": response.content.decode()}
        except Exception as e:
            response_data = {"error": str(e)}
    return render(request, 'core/test_endpoint.html', {'response': response_data})

# Nuevas vistas para los endpoints de API
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction
from .models import Carrito, ItemCarrito, Orden, ItemOrden, User, Producto
from .serializers import OrdenSerializer
from decimal import Decimal

class CrearOrdenView(APIView):
    def post(self, request):
        print("Datos recibidos:", request.data)
        try:
            usuario = request.user if request.user.is_authenticated else User.objects.get_or_create(username='testuser')[0]
        except User.DoesNotExist:
            usuario = User.objects.create(username='testuser')
            usuario.save()

        carrito = Carrito.objects.filter(
            usuario=usuario,
            creado__gte=timezone.now() - timezone.timedelta(minutes=15)
        ).order_by('-creado').first()
        if not carrito:
            print("No se encontró carrito válido")
            return Response({"error": "No se encontró un carrito válido"}, status=status.HTTP_400_BAD_REQUEST)

        items_carrito = ItemCarrito.objects.filter(carrito=carrito)
        if not items_carrito.exists():
            print("Carrito vacío")
            return Response({"error": "El carrito está vacío"}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data
        metodo_pago = data.get('metodo_pago', 'transferencia')
        total = Decimal(str(data.get('total', '0')))

        calculated_total = sum(item.cantidad * item.producto.precio for item in items_carrito)
        print(f"Calculated total: {calculated_total}, Received total: {total}")
        if abs(calculated_total - total) > Decimal('0.01'):
            print("Totales no coinciden")
            return Response({"error": "El total no coincide con los ítems"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            orden = Orden.objects.create(
                usuario=usuario,
                total=total,
                estado='confirmacion',
                metodo_pago=metodo_pago
            )
            for item in items_carrito:
                ItemOrden.objects.create(
                    orden=orden,
                    producto=item.producto,
                    cantidad=item.cantidad,
                    precio=item.producto.precio
                )
            items_carrito.delete()

        serializer = OrdenSerializer(orden)
        whatsapp_link = serializer.data['whatsapp_link']
        print(f"WhatsApp link: {whatsapp_link}")

        return Response({"whatsapp_link": whatsapp_link}, status=status.HTTP_200_OK)
    


from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Orden
from django.core.mail import send_mail
from django.conf import settings
from openpyxl.styles import Font, Alignment, PatternFill

def is_superuser(user):
    return user.is_superuser

def test_correo(request):
    try:
        from django.core.mail import send_mail
        send_mail(
            subject="PRUEBA - CORREO DESDE DISTRIBUIDORA",
            message="Si ves esto, ¡EL CORREO FUNCIONA PERFECTO!",
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=['valeveraleiva@gmail.com'],  # ← TU CORREO
            fail_silently=False,
        )
        return HttpResponse("<h1>CORREO ENVIADO. REVISA TU BANDEJA (y spam)</h1>")
    except Exception as e:
        return HttpResponse(f"<h1>ERROR: {e}</h1>")
    
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

@login_required
@user_passes_test(lambda u: u.is_superuser)
def gestion_estados(request):
    busqueda = request.GET.get('q', '').strip()
    estado_filtro = request.GET.get('estado', '')

    # Órdenes para la tabla principal - agregamos direccion_envio al select_related
    ordenes = Orden.objects.select_related('usuario', 'usuario__perfil', 'direccion_envio') \
                           .prefetch_related('itemorden_set__producto') \
                           .order_by('-fecha')

    if busqueda:
        ordenes = ordenes.filter(
            Q(id__icontains=busqueda) |
            Q(usuario__username__icontains=busqueda) |
            Q(usuario__perfil__nombre__icontains=busqueda) |
            Q(usuario__perfil__apellido_paterno__icontains=busqueda)
        )

    if estado_filtro:
        ordenes = ordenes.filter(estado=estado_filtro)

    # Cambio de estado
    if request.method == 'POST':
        orden_id = request.POST.get('orden_id')
        nuevo_estado = request.POST.get('estado')
        try:
            orden = Orden.objects.get(id=orden_id)
            estado_anterior = orden.get_estado_display()
            orden.estado = nuevo_estado
            orden.save()

            mensaje_wa = f"Hola! Mi pedido es el #{orden.id} - Estado: {orden.get_estado_display()}"
            whatsapp_link = f"https://wa.me/56949071013?text={urllib.parse.quote(mensaje_wa)}"

            if orden.usuario.email:
                try:
                    html_email = render_to_string('emails/cambio_estado.html', {
                        'cliente': orden.usuario.perfil.nombre_completo() if hasattr(orden.usuario, 'perfil') else orden.usuario.username,
                        'pedido_id': orden.id,
                        'estado_anterior': estado_anterior,
                        'nuevo_estado': orden.get_estado_display(),
                        'total': orden.total,
                        'whatsapp_link': whatsapp_link,
                        'items': orden.itemorden_set.all(),
                    })
                    email = EmailMultiAlternatives(
                        subject=f"¡Tu pedido #{orden.id} ha cambiado de estado!",
                        body="Tu pedido ha cambiado de estado.",
                        from_email=settings.EMAIL_HOST_USER,
                        to=[orden.usuario.email]
                    )
                    email.attach_alternative(html_email, "text/html")
                    email.send()
                except Exception as e:
                    print(f"ERROR ENVÍO CORREO: {e}")

            return JsonResponse({'success': True, 'nuevo_estado': orden.get_estado_display()})
        except Exception as e:
            print(f"ERROR CAMBIO ESTADO: {e}")
            return JsonResponse({'success': False})

    # Paginación
    paginator = Paginator(ordenes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Ventas diarias y mensuales (siempre calculamos)
    todas_ordenes = Orden.objects.all().order_by('-fecha')
    ventas_por_dia = {}
    for orden in todas_ordenes:
        dia = orden.fecha.date()
        ventas_por_dia.setdefault(dia, []).append(orden)

    dias_ordenados = sorted(ventas_por_dia.keys(), reverse=True)[:30]
    ventas_diarias_agrupadas = [(dia, ventas_por_dia[dia]) for dia in dias_ordenados]

    ventas_mensuales = todas_ordenes.extra({
        'mes': "strftime('%%m', fecha)",
        'año': "strftime('%%Y', fecha)"
    }).values('mes', 'año').annotate(
        total_ventas=Sum('total'),
        num_pedidos=Count('id')
    ).order_by('-año', '-mes')[:12]

    meses_nombres = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
    }
    for item in ventas_mensuales:
        item['nombre_mes'] = meses_nombres.get(item['mes'], item['mes'])

    # Exportar a Excel (protegido contra errores)
    if request.GET.get('export') == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Ventas"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="748C28", end_color="748C28", fill_type="solid")
            title_font = Font(size=16, bold=True)
            align_center = Alignment(horizontal="center")

            ws['A1'] = "REPORTE DE VENTAS - DISTRIBUIDORA TALAGANTE"
            ws.merge_cells('A1:E1')
            ws['A1'].font = title_font
            ws['A1'].alignment = align_center

            row = 3
            ws.cell(row=row, column=1, value="Ventas Diarias (Últimos 30 días)").font = Font(bold=True, size=14)
            row += 2

            headers = ["Fecha", "ID Pedido", "Cliente", "Total", "Estado"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center

            row += 1
            for dia, ordenes_dia in ventas_diarias_agrupadas:
                total_dia = sum(o.total for o in ordenes_dia)
                ws.cell(row=row, column=1, value=dia.strftime("%d/%m/%Y")).font = Font(bold=True)
                row += 1
                for o in ordenes_dia:
                    cliente = o.usuario.username
                    if hasattr(o.usuario, 'perfil') and o.usuario.perfil:
                        cliente = o.usuario.perfil.nombre_completo() or cliente
                    ws.append(["", f"#{o.id}", cliente, float(o.total), o.get_estado_display()])
                ws.cell(row=row, column=3, value="Total del día:").font = Font(bold=True)
                ws.cell(row=row, column=4, value=float(total_dia)).font = Font(bold=True, color="008000")
                row += 2

            row += 2
            ws.cell(row=row, column=1, value="Resumen Mensual").font = Font(bold=True, size=14)
            row += 2

            headers_mes = ["Mes", "N° Pedidos", "Total Ventas"]
            for col, h in enumerate(headers_mes, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            row += 1
            for mes in ventas_mensuales:
                ws.append([f"{mes['nombre_mes']} {mes['año']}", mes['num_pedidos'], float(mes['total_ventas'] or 0)])

            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 22

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'
            wb.save(response)
            return response

        except Exception as e:
            # Si falla el export, mostramos página normal con mensaje
            print(f"ERROR EXPORT EXCEL: {e}")
            messages.error(request, "Error al generar Excel. Intenta de nuevo más tarde.")
            # Continuamos con la vista normal

    context = {
        'ordenes': page_obj,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
        'estados_choices': Orden.ESTADOS,
        'ventas_diarias_agrupadas': ventas_diarias_agrupadas,
        'ventas_mensuales': ventas_mensuales,
    }
    return render(request, 'core/gestion_estados.html', context)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def pedidos_despacho(request):
    ordenes = Orden.objects.select_related('usuario', 'usuario__perfil') \
                           .prefetch_related('itemorden_set__producto') \
                           .filter(estado__in=['despacho', 'preparacion']) \
                           .order_by('-fecha')
    
    context = {'ordenes': ordenes, 'titulo': 'Pedidos en Despacho / Listos para Retiro'}
    return render(request, 'core/pedidos_despacho.html', context)  # ← NOMBRE CORRECTO


@login_required
@user_passes_test(lambda u: u.is_superuser)
def pedidos_finalizados(request):
    ordenes = Orden.objects.select_related('usuario', 'usuario__perfil') \
                           .prefetch_related('itemorden_set__producto') \
                           .filter(estado__in=['completado', 'cancelado']) \
                           .order_by('-fecha')
    
    context = {'ordenes': ordenes, 'titulo': 'Pedidos Finalizados'}
    return render(request, 'core/pedidos_finalizados.html', context)  # ← NOMBRE CORRECTO

@login_required
@user_passes_test(is_superuser)
def update_orden_status(request, orden_id):
    orden = get_object_or_404(Orden, id=orden_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in dict(Orden.ESTADOS):
            orden.estado = nuevo_estado
            orden.save()
            messages.success(request, f'Estado cambiado a {orden.get_estado_display()}')
        else:
            messages.error(request, 'Estado inválido')
        
        return redirect('/panel/')   
        
    context = {'orden': orden}
    return render(request, 'core/update_orden_status.html', context)


#Verificación correo electronico seguro
def enviar_codigo_verificacion(user):
    codigo_obj, _ = CodigoVerificacion.objects.get_or_create(usuario=user)
    codigo_obj.codigo = f"{random.randint(100000, 999999)}"
    codigo_obj.creado_en = timezone.now()
    codigo_obj.expirado = False
    codigo_obj.save()

    send_mail(
        "Tu código de verificación - Distribuidora Talagante",
        f"Hola {user.username}!\n\nTu código es:\n\n{codigo_obj.codigo}\n\nVálido por 10 minutos.\n\n¡Gracias!",
        None,
        [user.email],
        fail_silently=False,
    )

from django.http import JsonResponse
from .models import Producto

def api_buscar_por_codigo(request):
    codigo = request.GET.get('codigo', '').strip()
    
    if not codigo:
        return JsonResponse({'error': 'Código vacío'}, status=400)

    try:
        producto = Producto.objects.get(codigo_barra=codigo)
        return JsonResponse({
            'encontrado': True,
            'id': producto.id,
            'nombre': producto.nombre,
            'stock_actual': producto.stock,
            'mensaje': 'Producto encontrado. ¿Agregar stock?'
        })
    except Producto.DoesNotExist:
        return JsonResponse({
            'encontrado': False,
            'codigo_escaneado': codigo,
            'mensaje': 'Producto nuevo. Redirigiendo a creación...'
        })
    

@login_required
@user_passes_test(lambda u: u.is_superuser)
def escaneo_rapido(request):
    try:
        if request.method == "POST":
            codigo = request.POST.get("codigo_barras", "").strip()
            accion = request.POST.get("accion")

            if not codigo:
                messages.error(request, "No se recibió código de barras")
                return redirect('escaneo_rapido')

            try:
                producto = Producto.objects.get(codigo_barras=codigo)
                
                if accion == "confirmar":
                    try:
                        cantidad = Decimal(request.POST.get("cantidad", "0"))
                        if cantidad == 0:
                            messages.warning(request, "Ingresa una cantidad válida")
                        elif cantidad > 0:
                            producto.agregar_stock(cantidad)
                            messages.success(request, f"Stock sumado: +{cantidad} → {producto.stock} {producto.unidad_medida}")
                        else:
                            producto.restar_stock(abs(cantidad))
                            messages.success(request, f"Stock restado: -{abs(cantidad)} → {producto.stock} {producto.unidad_medida}")
                    except Exception as e:
                        messages.error(request, f"Error al actualizar stock: {e}")
                    return redirect('escaneo_rapido')

                return render(request, 'core/escaneo_existente.html', {
                    'producto': producto,
                    'codigo': codigo
                })

            except Producto.DoesNotExist:
                return redirect(f"{reverse('producto_create')}?codigo_barras={codigo}")

        # GET normal
        return render(request, 'core/escaneo_rapido.html')

    except Exception as e:
        # Esto nos muestra el error real en producción
        return HttpResponse(f"<pre>Error en escaneo_rapido: {type(e).__name__}: {str(e)}</pre>", status=500)


@staff_member_required
def redirigir_crear_producto_con_codigo(request):
    codigo = request.GET.get('codigo_barras', '').strip()
    if codigo:
        return redirect(f"{reverse('admin:core_producto_add')}?codigo_barras={codigo}")
    return redirect('admin:core_producto_add')

@csrf_exempt
def autocompletar_direccion(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        data = json.loads(request.body)
        q = data.get("q", "").strip()
        comuna = data.get("comuna", "").strip()

        if len(q) < 3 or not comuna:
            return JsonResponse({"sugerencias": []})

        headers = {
            "User-Agent": "DistribuidoraTalagante/1.0 (+56958530495)",
            "Accept-Language": "es-cl"
        }

        params = {
            "q": f"{q}, {comuna}, Región Metropolitana, Chile",
            "format": "json",
            "limit": 12,
            "countrycodes": "cl",
            "addressdetails": 1
        }

        response = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params=params,
            headers=headers,
            timeout=10
        )

        if response.status_code != 200:
            return JsonResponse({"sugerencias": []})

        resultados = response.json()
        sugerencias = []

        for item in resultados:
            display = item.get("display_name", "")
            if not display:
                continue

            calle = display.split(",", 1)[0].strip()

            if (calle.lower().startswith(('calle ', 'avenida ', 'pasaje ', 'camino ')) or 
                re.match(r'^[A-Za-z]', calle)):
                if q.lower() in calle.lower():
                    sugerencias.append({
                        "value": calle,
                        "label": calle
                    })

        # Eliminar duplicados
        visto = set()
        unicas = []
        for s in sugerencias:
            if s["value"].lower() not in visto:
                visto.add(s["value"].lower())
                unicas.append(s)
                if len(unicas) >= 8:
                    break

        return JsonResponse({"sugerencias": unicas})

    except Exception as e:
        print("ERROR AUTOCOMPLETAR:", e)
        return JsonResponse({"sugerencias": [], "debug": str(e)})

# --- VISTA PARA GESTIONAR BANNERS (FRONTEND) ---
def gestion_banners(request):
    
    if not request.user.is_staff:
        return redirect('home')

    
    if request.method == 'POST' and 'crear_banner' in request.POST:
        titulo = request.POST.get('titulo')
        imagen = request.FILES.get('imagen')
        if titulo and imagen:
            Banner.objects.create(titulo=titulo, imagen=imagen)
            return redirect('gestion_banners') 

    
    if request.method == 'POST' and 'eliminar_banner' in request.POST:
        banner_id = request.POST.get('banner_id')
        banner = get_object_or_404(Banner, id=banner_id)
        banner.delete()
        return redirect('gestion_banners')

    
    banners = Banner.objects.all().order_by('orden')
    return render(request, 'core/gestion_banners.html', {'banners': banners})
